from datetime import datetime, timezone
from io import BytesIO

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from app.models.checkout_section import CheckoutSection, CheckoutSectionType
from app.models.counter import Counter, CounterType
from app.models.store import Store
from app.models.trial_studio import TrialStudio, TrialStudioType
from app.models.trial_zone import TrialZone, TrialZoneGender, TrialZoneType
from app.services.ml_excel_service import CHECKOUT_COLUMNS, TRAINING_SHEET, TRIAL_COLUMNS, MLExcelService


class FakeCheckoutRepository:
    def __init__(self):
        self.store = Store(id=1, store_number="S-1", name="Store One", is_active=True)
        self.section = CheckoutSection(id=10, store_id=1, name="Main", section_type=CheckoutSectionType.REGULAR, is_active=True)
        self.counter = Counter(id=20, section_id=10, name="Counter 1", counter_type=CounterType.REGULAR, is_active=True, next_available_time=datetime.now(timezone.utc))

    def get_store_by_id(self, store_id): return self.store if store_id == 1 else None
    def list_sections_for_training(self, store_id): return [self.section]
    def list_counters_for_training(self, store_id): return [self.counter]
    def get_section_for_training(self, section_id): return self.section if section_id == 10 else None
    def get_counter_for_training(self, counter_id): return self.counter if counter_id == 20 else None
    def get_store_timezone(self, store_id): return "Asia/Kolkata"


class FakeTrialRepository:
    def __init__(self):
        self.store = Store(id=1, store_number="S-1", name="Store One", is_active=True)
        self.zone = TrialZone(id=30, store_id=1, name="Women", zone_type=TrialZoneType.REGULAR, gender=TrialZoneGender.FEMALE, is_active=True)
        self.studio = TrialStudio(id=40, trial_zone_id=30, name="Studio 1", studio_type=TrialStudioType.REGULAR, is_active=True, next_available_time=datetime.now(timezone.utc))

    def get_store(self, store_id): return self.store if store_id == 1 else None
    def list_zones(self, include_inactive, store_id): return [self.zone]
    def list_studios(self, include_inactive, store_id): return [self.studio]
    def get_zone(self, zone_id): return self.zone if zone_id == 30 else None
    def get_studio(self, studio_id): return self.studio if studio_id == 40 else None
    def get_trial_store_timezone(self, store_id): return "Asia/Kolkata"


@pytest.mark.parametrize(
    ("service", "columns"),
    [(MLExcelService(FakeCheckoutRepository(), "checkout"), CHECKOUT_COLUMNS), (MLExcelService(FakeTrialRepository(), "trial"), TRIAL_COLUMNS)],
)
def test_template_contains_required_sheets_and_headers(service, columns):
    workbook = load_workbook(BytesIO(service.build_template(1)), data_only=True)
    assert {"Instructions", TRAINING_SHEET, "Store Lookups"}.issubset(workbook.sheetnames)
    assert [cell.value for cell in workbook[TRAINING_SHEET][1]] == columns


def test_checkout_parser_uses_only_workbook_rows(monkeypatch):
    monkeypatch.setattr("app.services.ml_excel_service.settings.ML_MIN_TRAINING_SAMPLES", 2)
    service = MLExcelService(FakeCheckoutRepository(), "checkout")
    content = service.build_template(1)
    workbook = load_workbook(BytesIO(content))
    sheet = workbook[TRAINING_SHEET]
    sheet.append([sheet.cell(2, index).value for index in range(1, len(CHECKOUT_COLUMNS) + 1)])
    output = BytesIO(); workbook.save(output)

    rows = service.parse(1, output.getvalue())

    assert len(rows) == 2
    assert rows[0]["features"]["section_id"] == "10"
    assert rows[0]["features"]["assigned_counter_id"] == "20"


def test_invalid_cross_store_reference_rejects_entire_workbook(monkeypatch):
    monkeypatch.setattr("app.services.ml_excel_service.settings.ML_MIN_TRAINING_SAMPLES", 1)
    service = MLExcelService(FakeTrialRepository(), "trial")
    workbook = load_workbook(BytesIO(service.build_template(1)))
    sheet = workbook[TRAINING_SHEET]
    sheet.cell(2, TRIAL_COLUMNS.index("trial_zone_id") + 1, 999)
    output = BytesIO(); workbook.save(output)

    with pytest.raises(HTTPException) as exc_info:
        service.parse(1, output.getvalue())

    assert exc_info.value.status_code == 422
    assert any(error["column"] == "trial_zone_id" for error in exc_info.value.detail["errors"])
