from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Literal
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.core.config import settings


TRAINING_SHEET = "Training Data"
HEADER_FILL = PatternFill("solid", fgColor="6785B5")
CHECKOUT_COLUMNS = [
    "joined_at", "service_minutes", "item_count", "section_id", "assigned_counter_id",
    "section_busy_count_at_join", "section_active_counter_count_at_join",
    "recent_cancellation_rate", "recent_average_service_minutes", "promotion_day_flag",
    "basket_size", "cart_type", "customer_type",
]
TRIAL_COLUMNS = [
    "joined_at", "service_minutes", "item_count", "trial_zone_id", "assigned_studio_id",
    "trial_zone_busy_count_at_join", "trial_active_studio_count_at_join",
    "recent_cancellation_rate", "recent_average_service_minutes", "promotion_day_flag", "customer_type",
]


class MLExcelService:
    def __init__(self, repository: Any, module: Literal["checkout", "trial"]) -> None:
        self.repository = repository
        self.module = module

    def build_template(self, store_id: int) -> bytes:
        store = self._get_store(store_id)
        if store is None:
            raise HTTPException(status_code=404, detail="Store not found")
        wb = Workbook()
        instructions = wb.active
        instructions.title = "Instructions"
        instructions.append([f"{self.module.title()} ML training template"])
        instructions.append([f"Store: {store.name} ({store.store_number}), ID {store.id}"])
        instructions.append(["Replace the example row, then enter one completed service per row. Do not rename sheets or headers."])
        instructions.append([f"At least {settings.ML_MIN_TRAINING_SAMPLES} rows are required. Rates use decimals from 0 to 1."])
        instructions.append(["joined_at must be an Excel date/time or ISO date/time; service minutes must be between 1 and 240."])
        instructions["A1"].font = Font(bold=True, size=14)

        columns = CHECKOUT_COLUMNS if self.module == "checkout" else TRIAL_COLUMNS
        sheet = wb.create_sheet(TRAINING_SHEET)
        sheet.append(columns)
        example = self._example_row(store_id)
        sheet.append([example.get(column) for column in columns])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(columns)).coordinate}"
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(color="FFFFFF", bold=True)
        for index, column in enumerate(columns, 1):
            sheet.column_dimensions[sheet.cell(1, index).column_letter].width = max(16, len(column) + 2)

        if self.module == "checkout":
            sections = self.repository.list_sections_for_training(store_id)
            counters = self.repository.list_counters_for_training(store_id)
            lookup = wb.create_sheet("Store Lookups")
            lookup.append(["section_id", "section_name", "counter_id", "counter_name", "counter_section_id"])
            for idx in range(max(len(sections), len(counters))):
                section = sections[idx] if idx < len(sections) else None
                counter = counters[idx] if idx < len(counters) else None
                lookup.append([
                    section.id if section else None, section.name if section else None,
                    counter.id if counter else None, (counter.name or f"Counter {counter.id}") if counter else None,
                    counter.section_id if counter else None,
                ])
        else:
            zones = self.repository.list_zones(include_inactive=True, store_id=store_id)
            studios = self.repository.list_studios(include_inactive=True, store_id=store_id)
            lookup = wb.create_sheet("Store Lookups")
            lookup.append(["trial_zone_id", "zone_name", "zone_type", "gender", "studio_id", "studio_name", "studio_zone_id", "studio_type"])
            for idx in range(max(len(zones), len(studios))):
                zone = zones[idx] if idx < len(zones) else None
                studio = studios[idx] if idx < len(studios) else None
                lookup.append([
                    zone.id if zone else None, zone.name if zone else None,
                    zone.zone_type.value if zone else None, zone.gender.value if zone else None,
                    studio.id if studio else None, (studio.name or f"Studio {studio.id}") if studio else None,
                    studio.trial_zone_id if studio else None, studio.studio_type.value if studio else None,
                ])
        for cell in lookup[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(color="FFFFFF", bold=True)
        flag_validation = DataValidation(type="list", formula1='"0,1"')
        sheet.add_data_validation(flag_validation)
        flag_column = columns.index("promotion_day_flag") + 1
        flag_validation.add(f"{sheet.cell(2, flag_column).column_letter}2:{sheet.cell(2, flag_column).column_letter}10001")
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def parse(self, store_id: int, content: bytes) -> list[dict[str, object]]:
        if self._get_store(store_id) is None:
            raise HTTPException(status_code=404, detail="Store not found")
        try:
            wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception:
            self._reject([self._error(None, None, "Workbook is not a valid .xlsx file")])
        if TRAINING_SHEET not in wb.sheetnames:
            self._reject([self._error(None, None, f"Required sheet '{TRAINING_SHEET}' is missing")])
        sheet = wb[TRAINING_SHEET]
        expected = CHECKOUT_COLUMNS if self.module == "checkout" else TRIAL_COLUMNS
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        duplicates = sorted({header for header in headers if header and headers.count(header) > 1})
        errors = []
        if duplicates:
            errors.append(self._error(1, None, f"Duplicate headers: {', '.join(duplicates)}"))
        for column in expected:
            if column not in headers:
                errors.append(self._error(1, column, "Required column is missing"))
        unknown = [header for header in headers if header and header not in expected]
        if unknown:
            errors.append(self._error(1, None, f"Unknown columns: {', '.join(unknown)}"))
        if errors:
            self._reject(errors)

        rows = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            record = dict(zip(headers, values))
            if all(value is None or str(value).strip() == "" for value in record.values()):
                continue
            if len(rows) >= settings.ML_TRAINING_UPLOAD_MAX_ROWS:
                self._reject([self._error(row_number, None, f"Maximum {settings.ML_TRAINING_UPLOAD_MAX_ROWS} data rows allowed")])
            parsed, row_errors = self._parse_row(record, row_number, store_id)
            errors.extend(row_errors)
            if not row_errors:
                rows.append(parsed)
        if len(rows) < settings.ML_MIN_TRAINING_SAMPLES and not errors:
            errors.append(self._error(None, None, f"At least {settings.ML_MIN_TRAINING_SAMPLES} training rows are required"))
        if errors:
            self._reject(errors)
        return rows

    def _parse_row(self, record: dict[str, object], row: int, store_id: int):
        errors = []
        joined_at = self._datetime(record["joined_at"], row, "joined_at", errors)
        duration = self._number(record["service_minutes"], row, "service_minutes", errors, 1, 240)
        item_count = self._number(record["item_count"], row, "item_count", errors, 0, None)
        cancel_rate = self._number(record["recent_cancellation_rate"], row, "recent_cancellation_rate", errors, 0, 1)
        recent_average = self._number(record["recent_average_service_minutes"], row, "recent_average_service_minutes", errors, 0, 240)
        promotion = self._number(record["promotion_day_flag"], row, "promotion_day_flag", errors, 0, 1, integer=True)
        features: dict[str, object] = {
            "item_count": item_count, "recent_cancellation_rate": cancel_rate,
            "recent_average_service_minutes": recent_average, "promotion_day_flag": promotion,
        }
        if joined_at:
            local = joined_at.astimezone(self._store_timezone(store_id))
            features.update(hour_of_day=float(local.hour), day_of_week=float(local.weekday()), is_weekend=1.0 if local.weekday() >= 5 else 0.0)
        if self.module == "checkout":
            section_id = self._integer(record["section_id"], row, "section_id", errors)
            counter_id = self._integer(record["assigned_counter_id"], row, "assigned_counter_id", errors)
            section = self.repository.get_section_for_training(section_id) if section_id else None
            counter = self.repository.get_counter_for_training(counter_id) if counter_id else None
            if section is None or section.store_id != store_id:
                errors.append(self._error(row, "section_id", "Section does not belong to the selected store"))
            if counter is None or counter.section_id != section_id:
                errors.append(self._error(row, "assigned_counter_id", "Counter does not belong to the selected section"))
            features.update(
                section_busy_count_at_join=self._number(record["section_busy_count_at_join"], row, "section_busy_count_at_join", errors, 0, None),
                section_active_counter_count_at_join=self._number(record["section_active_counter_count_at_join"], row, "section_active_counter_count_at_join", errors, 0, None),
                basket_size=self._text(record["basket_size"], row, "basket_size", errors),
                cart_type=self._text(record["cart_type"], row, "cart_type", errors),
                customer_type=self._text(record["customer_type"], row, "customer_type", errors),
                section_id=str(section_id), assigned_counter_id=str(counter_id),
            )
        else:
            zone_id = self._integer(record["trial_zone_id"], row, "trial_zone_id", errors)
            studio_id = self._integer(record["assigned_studio_id"], row, "assigned_studio_id", errors)
            zone = self.repository.get_zone(zone_id) if zone_id else None
            studio = self.repository.get_studio(studio_id) if studio_id else None
            if zone is None or zone.store_id != store_id:
                errors.append(self._error(row, "trial_zone_id", "Trial zone does not belong to the selected store"))
            if studio is None or studio.trial_zone_id != zone_id:
                errors.append(self._error(row, "assigned_studio_id", "Studio does not belong to the selected trial zone"))
            features.update(
                trial_zone_busy_count_at_join=self._number(record["trial_zone_busy_count_at_join"], row, "trial_zone_busy_count_at_join", errors, 0, None),
                trial_active_studio_count_at_join=self._number(record["trial_active_studio_count_at_join"], row, "trial_active_studio_count_at_join", errors, 0, None),
                customer_type=self._text(record["customer_type"], row, "customer_type", errors),
                trial_zone_id=str(zone_id), assigned_studio_id=str(studio_id),
                trial_zone_type=zone.zone_type.value.lower() if zone else "unknown",
                trial_zone_gender=zone.gender.value.lower() if zone else "unknown",
                studio_type=studio.studio_type.value.lower() if studio else "unknown",
            )
        return {"features": features, "duration_minutes": duration}, errors

    def _get_store(self, store_id):
        return self.repository.get_store_by_id(store_id) if self.module == "checkout" else self.repository.get_store(store_id)

    def _example_row(self, store_id):
        now = datetime.now().replace(microsecond=0)
        if self.module == "checkout":
            sections = self.repository.list_sections_for_training(store_id)
            counters = self.repository.list_counters_for_training(store_id)
            section = sections[0] if sections else None
            counter = next((c for c in counters if section and c.section_id == section.id), None)
            return dict(joined_at=now, service_minutes=8, item_count=12, section_id=section.id if section else None,
                        assigned_counter_id=counter.id if counter else None, section_busy_count_at_join=3,
                        section_active_counter_count_at_join=2, recent_cancellation_rate=0.05,
                        recent_average_service_minutes=7.5, promotion_day_flag=0, basket_size="medium",
                        cart_type="basket", customer_type="regular")
        zones = self.repository.list_zones(include_inactive=True, store_id=store_id)
        studios = self.repository.list_studios(include_inactive=True, store_id=store_id)
        zone = zones[0] if zones else None
        studio = next((s for s in studios if zone and s.trial_zone_id == zone.id), None)
        return dict(joined_at=now, service_minutes=10, item_count=3, trial_zone_id=zone.id if zone else None,
                    assigned_studio_id=studio.id if studio else None, trial_zone_busy_count_at_join=2,
                    trial_active_studio_count_at_join=3, recent_cancellation_rate=0.04,
                    recent_average_service_minutes=9.5, promotion_day_flag=0, customer_type="regular")

    def _store_timezone(self, store_id):
        name = self.repository.get_store_timezone(store_id) if self.module == "checkout" else self.repository.get_trial_store_timezone(store_id)
        try:
            return ZoneInfo(name or "Asia/Kolkata")
        except ZoneInfoNotFoundError:
            return ZoneInfo("Asia/Kolkata")

    def _number(self, value, row, column, errors, minimum, maximum, integer=False):
        try:
            number = float(value)
            if integer and not number.is_integer(): raise ValueError
            if minimum is not None and number < minimum: raise ValueError
            if maximum is not None and number > maximum: raise ValueError
            return int(number) if integer else number
        except (TypeError, ValueError):
            range_text = f" between {minimum} and {maximum}" if maximum is not None else f" at least {minimum}"
            errors.append(self._error(row, column, f"Must be a number{range_text}"))
            return 0

    def _integer(self, value, row, column, errors):
        return int(self._number(value, row, column, errors, 1, None, integer=True))

    def _text(self, value, row, column, errors):
        text = str(value).strip().lower() if value is not None else ""
        if not text:
            errors.append(self._error(row, column, "Value is required"))
        return text or "unknown"

    def _datetime(self, value, row, column, errors):
        parsed = value if isinstance(value, datetime) else None
        if parsed is None and isinstance(value, str):
            try: parsed = datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
            except ValueError: pass
        if parsed is None:
            errors.append(self._error(row, column, "Must be an Excel or ISO date/time"))
            return None
        if parsed.tzinfo is None: parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)

    def _error(self, row, column, message):
        return {"sheet": TRAINING_SHEET, "row": row, "column": column, "message": message}

    def _reject(self, errors):
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail={"message": "Training workbook validation failed", "errors": errors[:200]})
